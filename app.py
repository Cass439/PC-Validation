import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="PC Validation", layout="wide")

st.title("PC Validation")
st.write("Upload a Partner Catalog Excel file and a sales CSV. Any GTIN matched to BARCODE with TOTAL_SALES = 0 will have Price set to 0.")

catalog_file = st.file_uploader("Upload Partner Catalog (.xlsx)", type=["xlsx"])
sales_file = st.file_uploader("Upload Sales CSV", type=["csv"])

def normalize(value):
    return str(value).strip().upper()

def find_header_row(sheet, required_headers):
    for row in sheet.iter_rows(min_row=1, max_row=10):
        values = [cell.value for cell in row]
        normalized = [normalize(v) if v is not None else "" for v in values]
        if all(header in normalized for header in required_headers):
            return row[0].row
    return None

if catalog_file and sales_file:
    try:
        sales_df = pd.read_csv(sales_file)
        sales_df.columns = [c.strip().upper() for c in sales_df.columns]

        if "BARCODE" not in sales_df.columns or "TOTAL_SALES" not in sales_df.columns:
            st.error("Sales CSV must contain BARCODE and TOTAL_SALES columns.")
            st.stop()

        zero_sales_barcodes = set(
            sales_df.loc[pd.to_numeric(sales_df["TOTAL_SALES"], errors="coerce").fillna(0) == 0, "BARCODE"]
            .astype(str)
            .str.strip()
        )

        wb = load_workbook(catalog_file)
        ws = wb[wb.sheetnames[0]]

        header_row_num = find_header_row(ws, ["GTIN", "PRICE"])
        if header_row_num is None:
            st.error("Could not find a header row with GTIN and Price in the Excel file.")
            st.stop()

        headers = {}
        for cell in ws[header_row_num]:
            if cell.value is not None:
                headers[normalize(cell.value)] = cell.column

        gtin_col = headers.get("GTIN")
        price_col = headers.get("PRICE")

        if not gtin_col or not price_col:
            st.error("Could not locate GTIN and Price columns.")
            st.stop()

        updated_count = 0

        for row in range(header_row_num + 1, ws.max_row + 1):
            gtin_value = ws.cell(row=row, column=gtin_col).value
            if gtin_value is None:
                continue

            gtin_str = str(gtin_value).strip()
            if gtin_str in zero_sales_barcodes:
                ws.cell(row=row, column=price_col).value = 0
                updated_count += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success(f"Done. Updated {updated_count} rows.")
        st.download_button(
            "Download updated Partner Catalog",
            data=output,
            file_name="updated_partner_catalog.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Error: {e}")
